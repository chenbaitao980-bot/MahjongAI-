from __future__ import annotations

import logging
from datetime import datetime
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

_LOGGER = logging.getLogger("mahjongai.excel_logger")

_MELD_CN = {
    "pon": "碰", "chi": "吃",
    "kan_open": "明杠", "kan_closed": "暗杠", "kan_added": "补杠",
}

# (列标题, 列宽)
_COLUMNS = [
    ("序号", 6), ("时间", 10), ("事件", 14), ("触发方", 8),
    ("涉及牌", 10), ("我方手牌", 40), ("我方弃牌", 30), ("我方副露", 22),
    ("对方弃牌", 30), ("对方副露", 22), ("剩余牌数", 8), ("数据来源", 8),
    ("识别正确?", 12), ("纠正值", 22), ("备注", 26),
]

_FILL_HDR  = "2C3E50"   # 标题行背景（深灰蓝）
_FILL_DEAL = "D5F5E3"   # 开局行（浅绿）
_FILL_WIN  = "FDEBD0"   # 胡牌行（浅橙）
_FILL_SELF = "EBF5FB"   # 我方区列 F G H（浅蓝）
_FILL_OPP  = "FEF0E7"   # 对方区列 I J（浅橙）
_FILL_FIX  = "FFFACD"   # 纠错列 M N（浅黄）

# 列编号（1-based）
_COL_SELF_RANGE = (6, 7, 8)    # F G H
_COL_OPP_RANGE  = (9, 10)      # I J
_COL_FIX_RANGE  = (13, 14)     # M N


def _cn_tiles(tile_ids: list[str]) -> str:
    from game.tiles import tile_display_name
    return " ".join(tile_display_name(t) for t in tile_ids if t)


def _cn_melds(melds: list[dict[str, Any]]) -> str:
    from game.tiles import tile_display_name
    parts: list[str] = []
    for m in melds:
        cn = _MELD_CN.get(str(m.get("type", "")), "副露")
        tiles_cn = "".join(tile_display_name(t) for t in m.get("tiles", []))
        parts.append(f"{cn}[{tiles_cn}]")
    return "  ".join(parts)


def _tile_cn(tile: str) -> str:
    """tile_id（如'3m'）→ 中文名；已是中文则直接返回。"""
    if not tile:
        return ""
    if len(tile) <= 3 and tile[-1:] in "mpsz":
        from game.tiles import tile_display_name
        return tile_display_name(tile)
    return tile


class ExcelGameLogger:
    """把 npcap 每次牌面变更追加到 xlsx 文件。"""

    def __init__(self, path: str, flush_every: int = 5) -> None:
        self._path = path
        self._closed = False
        self._seq = 0
        self._flush_every = max(1, int(flush_every))
        self._wb: Any = None
        self._ws: Any = None
        if not _HAS_OPENPYXL:
            _LOGGER.warning("ExcelGameLogger init skipped: openpyxl unavailable path=%s", path)
            return
        self._wb = openpyxl.Workbook()
        self._ws = self._wb.active
        self._ws.title = "牌面流水"
        self._init_sheet()
        _LOGGER.info("ExcelGameLogger created path=%s flush_every=%s", path, self._flush_every)

    def _init_sheet(self) -> None:
        ws = self._ws
        hdr_fill = PatternFill("solid", fgColor=_FILL_HDR)
        hdr_font = Font(color="FFFFFF", bold=True, size=11, name="微软雅黑")
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for i, (name, width) in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=1, column=i, value=name)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = hdr_align
            # 列宽用 column_dimensions key = 字母
            col_letter = ws.cell(row=1, column=i).column_letter
            ws.column_dimensions[col_letter].width = width
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"
        last_col_letter = ws.cell(row=1, column=len(_COLUMNS)).column_letter
        ws.auto_filter.ref = f"A1:{last_col_letter}1"
        # M 列下拉校验
        dv = DataValidation(
            type="list",
            formula1='"未确认,✓正确,✗错误"',
            allow_blank=False,
            showDropDown=False,
            showErrorMessage=True,
            error="请从下拉选择：未确认 / ✓正确 / ✗错误",
            errorTitle="输入无效",
        )
        dv.sqref = "M2:M20000"
        ws.add_data_validation(dv)

    def log_row(
        self,
        *,
        event_type: str,
        actor: str,
        tile: str,
        self_hand: list[str],
        self_discards: list[str],
        self_melds: list[dict],
        enemy_discards: list[str],
        enemy_melds: list[dict],
        remaining: int,
        source: str,
        ts: str = "",
    ) -> None:
        if self._closed or self._ws is None:
            return
        try:
            self._seq += 1
            row_ts = ts if ts else datetime.now().strftime("%H:%M:%S")
            row_data = [
                self._seq,                        # A 序号
                row_ts,                           # B 时间
                event_type,                       # C 事件
                actor,                            # D 触发方
                _tile_cn(tile),                   # E 涉及牌
                _cn_tiles(self_hand),             # F 我方手牌
                _cn_tiles(self_discards),         # G 我方弃牌
                _cn_melds(self_melds),            # H 我方副露
                _cn_tiles(enemy_discards),        # I 对方弃牌
                _cn_melds(enemy_melds),           # J 对方副露
                remaining,                        # K 剩余牌数
                source,                           # L 数据来源
                "未确认",                          # M 识别正确?
                "",                               # N 纠正值
                "",                               # O 备注
            ]
            row_idx = self._seq + 1
            is_deal = event_type in ("开局发牌", "开局标记")
            is_win  = event_type == "胡牌"
            ws = self._ws
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                if is_deal:
                    cell.fill = PatternFill("solid", fgColor=_FILL_DEAL)
                elif is_win:
                    cell.fill = PatternFill("solid", fgColor=_FILL_WIN)
                else:
                    if col_idx in _COL_SELF_RANGE:
                        cell.fill = PatternFill("solid", fgColor=_FILL_SELF)
                    elif col_idx in _COL_OPP_RANGE:
                        cell.fill = PatternFill("solid", fgColor=_FILL_OPP)
                    elif col_idx in _COL_FIX_RANGE:
                        cell.fill = PatternFill("solid", fgColor=_FILL_FIX)
            if self._seq % self._flush_every == 0:
                self._flush()
        except Exception as exc:
            _LOGGER.warning("ExcelGameLogger log_row failed seq=%s: %s", self._seq, exc)

    def _flush(self) -> None:
        if self._wb is None or self._closed:
            return
        try:
            self._wb.save(self._path)
            _LOGGER.info("ExcelGameLogger flushed seq=%s path=%s", self._seq, self._path)
        except Exception as exc:
            _LOGGER.warning("ExcelGameLogger flush failed seq=%s: %s", self._seq, exc)

    def close(self) -> str:
        if self._closed:
            return self._path
        self._closed = True
        saved = False
        if self._wb is not None:
            try:
                self._wb.save(self._path)
                saved = True
            except Exception as exc:
                _LOGGER.warning("ExcelGameLogger close save failed seq=%s: %s", self._seq, exc)
        _LOGGER.info("ExcelGameLogger closed path=%s seq=%s saved=%s", self._path, self._seq, saved)
        return self._path
